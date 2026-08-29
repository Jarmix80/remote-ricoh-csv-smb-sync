"""Cykliczny import raportow Documaster z SMB do Firebird CMAIL."""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import re
import sqlite3
import warnings
from collections import Counter
from collections.abc import Callable
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from pathlib import Path, PureWindowsPath
from typing import Any

from openpyxl import load_workbook

from .config import WARSAW_TZ
from .firebird_cmail import CounterRecord, DeviceMatch, FirebirdCmailImporter
from .smb_io import SmbClient

DEFAULT_DOCUMASTER_DB = Path("local/documaster/documaster.sqlite")
DEFAULT_DOCUMASTER_REPORT_DIR = Path("local/documaster/reports")
DOCUMASTER_SMB_DIR = "documaster"
DOCUMASTER_MIN_FILE_AGE_SECONDS = 5 * 60
DOCUMASTER_WRITE_GUARD = "DOCUMASTER_ALLOW_WRITES"
SUPPORTED_SUFFIXES = {".csv", ".xlsx"}
REQUIRED_HEADERS = {
    "Numer seryjny",
    "Data ostatniego odczytu",
    "Licznik główny",
    "Suma Cz/B",
    "Suma Kolor",
}


class DocumasterParseError(ValueError):
    """Plik Documaster ma nieobslugiwany albo niepoprawny format."""


@dataclass(frozen=True, slots=True)
class DocumasterRecord:
    """Znormalizowany pojedynczy odczyt z raportu Documaster."""

    row_number: int
    serial: str
    counter_datetime: dt.datetime
    source_model: str
    total: int
    mono: int
    color: int
    copier_total: int | None
    copier_mono: int | None
    copier_color: int | None
    printer_total: int | None
    printer_mono: int | None
    printer_color: int | None
    scanner_total: int | None

    def as_counter_record(self, device: DeviceMatch) -> CounterRecord:
        """Buduje rekord CMAIL z kanoniczna marka i modelem z MASZYNA."""
        return CounterRecord(
            serial=self.serial,
            counter_date=self.counter_datetime,
            brand=device.brand,
            model=device.model,
            total=self.total,
            mono=self.mono,
            color=self.color,
            copier_total=self.copier_total,
            copier_mono=self.copier_mono,
            copier_color=self.copier_color,
            printer_total=self.printer_total,
            printer_mono=self.printer_mono,
            printer_color=self.printer_color,
            scan_total=self.scanner_total,
        )


@dataclass(frozen=True, slots=True)
class ParsedDocumasterFile:
    """Poprawnie odczytany raport Documaster."""

    source_name: str
    report_customer: str
    records: list[DocumasterRecord]


@dataclass(frozen=True, slots=True)
class DocumasterRowResult:
    """Wynik kontroli albo importu jednego wiersza."""

    row_number: int
    serial: str
    counter_datetime: str
    total: int
    mono: int
    color: int
    status: str
    message: str = ""
    id_device: str = ""
    id_customer: str = ""
    id_cpc: str = ""
    expected_customer: str = ""
    latest_counter_date: str = ""
    cmail_id: str = ""

    def as_csv_row(self, file_result: DocumasterFileResult) -> dict[str, object]:
        return {
            "source_file": file_result.source_name,
            "file_hash": file_result.file_hash,
            "report_customer": file_result.report_customer,
            "file_status": file_result.status,
            "row_number": self.row_number,
            "serial": self.serial,
            "counter_datetime": self.counter_datetime,
            "total": self.total,
            "mono": self.mono,
            "color": self.color,
            "status": self.status,
            "id_device": self.id_device,
            "id_customer": self.id_customer,
            "id_cpc": self.id_cpc,
            "expected_customer": self.expected_customer,
            "latest_counter_date": self.latest_counter_date,
            "cmail_id": self.cmail_id,
            "message": self.message,
            "archive_path": file_result.archive_path,
        }


@dataclass(slots=True)
class DocumasterFileResult:
    """Wynik przetworzenia jednego pliku."""

    source_name: str
    file_hash: str = ""
    file_size: int = 0
    source_mtime: float = 0.0
    report_customer: str = ""
    status: str = ""
    message: str = ""
    archive_path: str = ""
    rows: list[DocumasterRowResult] = field(default_factory=list)

    @property
    def status_counts(self) -> Counter[str]:
        return Counter(row.status for row in self.rows)

    @property
    def has_warning(self) -> bool:
        return self.status in {"failed_invalid", "failed_retryable", "processed_warning"} or any(
            row.status.startswith("skipped_") for row in self.rows
        )


@dataclass(frozen=True, slots=True)
class DocumasterRunResult:
    """Podsumowanie pojedynczego skanu katalogu Documaster."""

    run_id: int
    execute: bool
    files: list[DocumasterFileResult]
    report_path: Path

    @property
    def status_counts(self) -> Counter[str]:
        counts: Counter[str] = Counter()
        for item in self.files:
            counts.update(item.status_counts)
        return counts

    @property
    def has_warning(self) -> bool:
        return any(item.has_warning for item in self.files)

    def as_log_message(self) -> str:
        file_counts = Counter(item.status for item in self.files)
        files = ", ".join(f"{key}={value}" for key, value in sorted(file_counts.items()))
        rows = ", ".join(f"{key}={value}" for key, value in sorted(self.status_counts.items()))
        return (
            f"Documaster: execute={self.execute}, files={len(self.files)}"
            f" ({files or 'brak'}), rows=({rows or 'brak'})."
        )


class DocumasterStore:
    """Lokalny rejestr uruchomien i przetworzonych plikow."""

    def __init__(self, db_path: Path = DEFAULT_DOCUMASTER_DB) -> None:
        self.db_path = db_path.expanduser().resolve()

    def initialize(self) -> None:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        with self._connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    started_at TEXT NOT NULL,
                    finished_at TEXT,
                    execute_mode INTEGER NOT NULL,
                    status TEXT NOT NULL,
                    report_path TEXT NOT NULL DEFAULT '',
                    message TEXT NOT NULL DEFAULT ''
                );

                CREATE TABLE IF NOT EXISTS files (
                    file_hash TEXT PRIMARY KEY,
                    source_name TEXT NOT NULL,
                    file_size INTEGER NOT NULL,
                    source_mtime REAL NOT NULL,
                    report_customer TEXT NOT NULL DEFAULT '',
                    status TEXT NOT NULL,
                    first_seen_at TEXT NOT NULL,
                    processed_at TEXT,
                    archive_path TEXT NOT NULL DEFAULT '',
                    report_path TEXT NOT NULL DEFAULT '',
                    row_count INTEGER NOT NULL DEFAULT 0,
                    inserted_count INTEGER NOT NULL DEFAULT 0,
                    duplicate_count INTEGER NOT NULL DEFAULT 0,
                    skipped_count INTEGER NOT NULL DEFAULT 0,
                    message TEXT NOT NULL DEFAULT ''
                );
                """
            )

    def start_run(self, execute: bool, now: dt.datetime) -> int:
        with self._connect() as connection:
            cursor = connection.execute(
                """
                INSERT INTO runs (started_at, execute_mode, status)
                VALUES (?, ?, 'running')
                """,
                (now.isoformat(), int(execute)),
            )
            return int(cursor.lastrowid)

    def finish_run(
        self,
        run_id: int,
        *,
        status: str,
        report_path: Path,
        message: str,
        now: dt.datetime,
    ) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                UPDATE runs
                SET finished_at = ?, status = ?, report_path = ?, message = ?
                WHERE id = ?
                """,
                (now.isoformat(), status, str(report_path.resolve()), message, run_id),
            )

    def get_file(self, file_hash: str) -> sqlite3.Row | None:
        with self._connect() as connection:
            return connection.execute(
                "SELECT * FROM files WHERE file_hash = ?",
                (file_hash,),
            ).fetchone()

    def record_file(
        self,
        result: DocumasterFileResult,
        report_path: Path,
        now: dt.datetime,
    ) -> None:
        if not result.file_hash:
            return
        counts = result.status_counts
        skipped = sum(count for status, count in counts.items() if status.startswith("skipped_"))
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO files (
                    file_hash, source_name, file_size, source_mtime, report_customer,
                    status, first_seen_at, processed_at, archive_path, report_path,
                    row_count, inserted_count, duplicate_count, skipped_count, message
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(file_hash) DO UPDATE SET
                    source_name = excluded.source_name,
                    file_size = excluded.file_size,
                    source_mtime = excluded.source_mtime,
                    report_customer = excluded.report_customer,
                    status = excluded.status,
                    processed_at = excluded.processed_at,
                    archive_path = excluded.archive_path,
                    report_path = excluded.report_path,
                    row_count = excluded.row_count,
                    inserted_count = excluded.inserted_count,
                    duplicate_count = excluded.duplicate_count,
                    skipped_count = excluded.skipped_count,
                    message = excluded.message
                """,
                (
                    result.file_hash,
                    result.source_name,
                    result.file_size,
                    result.source_mtime,
                    result.report_customer,
                    result.status,
                    now.isoformat(),
                    now.isoformat(),
                    result.archive_path,
                    str(report_path.resolve()),
                    len(result.rows),
                    counts["inserted"],
                    counts["duplicate"],
                    skipped,
                    result.message,
                ),
            )

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.db_path)
        connection.row_factory = sqlite3.Row
        return connection


class DocumasterFirebirdImporter:
    """Planuje i wykonuje transakcyjny import rekordow Documaster."""

    def __init__(self, importer: FirebirdCmailImporter) -> None:
        self.importer = importer

    def process(
        self,
        parsed: ParsedDocumasterFile,
        *,
        execute: bool,
        file_hash: str,
    ) -> list[DocumasterRowResult]:
        connection = self.importer._connect()  # noqa: SLF001
        try:
            cursor = connection.cursor()
            matches = {
                record.serial.casefold(): self._fetch_device_matches(cursor, record.serial)
                for record in parsed.records
            }
            expected_customer = _expected_customer_id(matches)
            rows: list[DocumasterRowResult] = []
            for record in parsed.records:
                rows.append(
                    self._process_record(
                        cursor,
                        record,
                        matches[record.serial.casefold()],
                        expected_customer=expected_customer,
                        execute=execute,
                        file_hash=file_hash,
                    )
                )
            if execute:
                connection.commit()
            return rows
        except Exception:
            if execute and hasattr(connection, "rollback"):
                connection.rollback()
            raise
        finally:
            connection.close()

    @staticmethod
    def _fetch_device_matches(cursor: Any, serial: str) -> list[DeviceMatch]:
        cursor.execute(
            """
            SELECT ID_MASZYNA, ID_KLIENT, ID_UMOWACPC,
                   TRIM(COALESCE(MARKA, '')), TRIM(COALESCE(MODEL, ''))
            FROM MASZYNA
            WHERE TRIM(COALESCE(SERIAL, '')) = ?
               OR TRIM(COALESCE(SERIAL2, '')) = ?
            ORDER BY ID_MASZYNA
            """,
            (serial, serial),
        )
        return [
            DeviceMatch(
                id_maszyna=int(row[0]),
                id_klient=int(row[1]),
                id_umowacpc=int(row[2]) if row[2] is not None else None,
                brand=row[3] or "",
                model=row[4] or "",
            )
            for row in cursor.fetchall()
        ]

    def _process_record(
        self,
        cursor: Any,
        record: DocumasterRecord,
        matches: list[DeviceMatch],
        *,
        expected_customer: int | None,
        execute: bool,
        file_hash: str,
    ) -> DocumasterRowResult:
        base = {
            "row_number": record.row_number,
            "serial": record.serial,
            "counter_datetime": record.counter_datetime.isoformat(sep=" "),
            "total": record.total,
            "mono": record.mono,
            "color": record.color,
            "expected_customer": str(expected_customer or ""),
        }
        if not matches:
            return DocumasterRowResult(
                **base,
                status="skipped_device_not_found",
                message="Brak urzadzenia w MASZYNA.",
            )
        if len(matches) > 1:
            return DocumasterRowResult(
                **base,
                status="skipped_device_ambiguous",
                message=f"Numer seryjny pasuje do {len(matches)} urzadzen.",
            )

        device = matches[0]
        device_fields = {
            "id_device": str(device.id_maszyna),
            "id_customer": str(device.id_klient),
            "id_cpc": str(device.id_umowacpc or ""),
        }
        counter_record = record.as_counter_record(device)
        if self.importer._has_duplicate(cursor, counter_record):  # noqa: SLF001
            return DocumasterRowResult(
                **base,
                **device_fields,
                status="duplicate",
                message="Identyczny odczyt jest juz w CMAIL.",
            )
        if expected_customer is None:
            return DocumasterRowResult(
                **base,
                **device_fields,
                status="skipped_customer_ambiguous",
                message="Nie mozna jednoznacznie ustalic klienta raportu.",
            )
        if device.id_klient != expected_customer:
            return DocumasterRowResult(
                **base,
                **device_fields,
                status="skipped_customer_mismatch",
                message=(
                    f"Urzadzenie ma ID_KLIENT={device.id_klient}, "
                    f"raport wskazuje ID_KLIENT={expected_customer}."
                ),
            )

        latest = self._fetch_latest_counter(cursor, record.serial)
        latest_date = latest[0] if latest else None
        if latest and latest_date and record.counter_datetime.date() >= latest_date:
            incoming = (record.total, record.mono, record.color)
            existing = tuple(int(value or 0) for value in latest[1:4])
            if any(new < old for new, old in zip(incoming, existing, strict=True)):
                return DocumasterRowResult(
                    **base,
                    **device_fields,
                    latest_counter_date=latest_date.isoformat(),
                    status="skipped_counter_regression",
                    message=f"Nowy odczyt jest nizszy od ostatniego: {existing}.",
                )

        latest_text = latest_date.isoformat() if latest_date else ""
        if not execute:
            return DocumasterRowResult(
                **base,
                **device_fields,
                latest_counter_date=latest_text,
                status="would_insert",
                message="Dry-run: rekord zostalby dodany do CMAIL.",
            )

        cmail_id = self.importer._insert_cmail(  # noqa: SLF001
            cursor,
            counter_record,
            device,
            mailfrom="[import] - Documaster automate",
            comments=f"{dt.date.today().isoformat()}_documaster:{file_hash[:12]}",
        )
        return DocumasterRowResult(
            **base,
            **device_fields,
            latest_counter_date=latest_text,
            cmail_id=str(cmail_id),
            status="inserted",
            message="Dodano rekord do CMAIL.",
        )

    @staticmethod
    def _fetch_latest_counter(cursor: Any, serial: str) -> tuple[Any, ...] | None:
        cursor.execute(
            """
            SELECT FIRST 1 COUNTER_DATE, TOTAL, TOTAL_MONO, TOTAL_COLOR
            FROM CMAIL
            WHERE TRIM(COALESCE(SERIAL, '')) = ?
            ORDER BY COUNTER_DATE DESC, DATEIN DESC, ID_CMAIL DESC
            """,
            (serial,),
        )
        return cursor.fetchone()


def parse_documaster_file(source_name: str, payload: bytes) -> ParsedDocumasterFile:
    """Parsuje raport CSV albo XLSX i zwraca ujednolicone rekordy."""
    suffix = Path(source_name).suffix.casefold()
    if suffix == ".csv":
        raw_rows = _read_csv_rows(payload)
    elif suffix == ".xlsx":
        raw_rows = _read_xlsx_rows(payload)
    else:
        raise DocumasterParseError(f"Nieobslugiwany format pliku: {source_name}")

    if len(raw_rows) < 3:
        raise DocumasterParseError("Raport nie zawiera naglowka i danych.")
    title = str(raw_rows[0][0] or "").strip()
    report_customer = _parse_report_customer(title)
    headers = [str(value or "").strip() for value in raw_rows[1]]
    missing = sorted(REQUIRED_HEADERS - set(headers))
    if missing:
        raise DocumasterParseError(f"Brak wymaganych kolumn: {', '.join(missing)}")

    records: list[DocumasterRecord] = []
    for row_number, values in enumerate(raw_rows[2:], start=3):
        if not any(str(value or "").strip() for value in values):
            continue
        row = dict(zip(headers, values, strict=False))
        records.append(_parse_documaster_row(row_number, row))
    if not records:
        raise DocumasterParseError("Raport nie zawiera rekordow licznikow.")
    return ParsedDocumasterFile(
        source_name=source_name,
        report_customer=report_customer,
        records=records,
    )


def run_documaster_scan(
    *,
    smb: SmbClient,
    importer: FirebirdCmailImporter,
    store: DocumasterStore,
    execute: bool,
    log: Callable[[str], None],
    now: dt.datetime | None = None,
) -> DocumasterRunResult:
    """Skanuje katalog SMB, importuje nowe pliki i zapisuje raport."""
    current = now or dt.datetime.now(tz=WARSAW_TZ)
    store.initialize()
    run_id = store.start_run(execute, current)
    results: list[DocumasterFileResult] = []
    report_path = DEFAULT_DOCUMASTER_REPORT_DIR / "pending.csv"
    try:
        entries = smb.list_directory([DOCUMASTER_SMB_DIR])
        source_names = [
            name for name in entries if Path(name).suffix.casefold() in SUPPORTED_SUFFIXES
        ]
        log(
            f"Documaster: start skanu, tryb={'EXECUTE' if execute else 'DRY-RUN'}, "
            f"pliki={len(source_names)}."
        )
        firebird = DocumasterFirebirdImporter(importer)
        for source_name in source_names:
            result = _process_source_file(
                smb=smb,
                firebird=firebird,
                store=store,
                source_name=source_name,
                execute=execute,
                now=current,
                log=log,
            )
            results.append(result)

        report_path = write_documaster_report(results, current)
        if execute:
            for result in results:
                if result.file_hash:
                    store.record_file(result, report_path, current)
        status = "warning" if any(item.has_warning for item in results) else "success"
        summary = DocumasterRunResult(run_id, execute, results, report_path).as_log_message()
        store.finish_run(
            run_id,
            status=status,
            report_path=report_path,
            message=summary,
            now=current,
        )
        return DocumasterRunResult(run_id, execute, results, report_path)
    except Exception as exc:
        store.finish_run(
            run_id,
            status="failed",
            report_path=report_path,
            message=f"{type(exc).__name__}: {exc}",
            now=current,
        )
        raise


def write_documaster_report(
    results: list[DocumasterFileResult],
    now: dt.datetime,
    report_dir: Path = DEFAULT_DOCUMASTER_REPORT_DIR,
) -> Path:
    """Zapisuje szczegolowy raport CSV z jednego skanu."""
    report_dir.mkdir(parents=True, exist_ok=True)
    path = report_dir / f"documaster_{now:%Y%m%d_%H%M%S_%f}.csv"
    fieldnames = [
        "source_file",
        "file_hash",
        "report_customer",
        "file_status",
        "row_number",
        "serial",
        "counter_datetime",
        "total",
        "mono",
        "color",
        "status",
        "id_device",
        "id_customer",
        "id_cpc",
        "expected_customer",
        "latest_counter_date",
        "cmail_id",
        "message",
        "archive_path",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            if result.rows:
                writer.writerows(row.as_csv_row(result) for row in result.rows)
            else:
                writer.writerow(
                    {
                        "source_file": result.source_name,
                        "file_hash": result.file_hash,
                        "report_customer": result.report_customer,
                        "file_status": result.status,
                        "status": result.status,
                        "message": result.message,
                        "archive_path": result.archive_path,
                    }
                )
    return path


def _process_source_file(
    *,
    smb: SmbClient,
    firebird: DocumasterFirebirdImporter,
    store: DocumasterStore,
    source_name: str,
    execute: bool,
    now: dt.datetime,
    log: Callable[[str], None],
) -> DocumasterFileResult:
    stat = smb.stat([DOCUMASTER_SMB_DIR, source_name])
    result = DocumasterFileResult(
        source_name=source_name,
        file_size=int(stat.st_size),
        source_mtime=float(stat.st_mtime),
    )
    age_seconds = now.timestamp() - result.source_mtime
    if age_seconds < DOCUMASTER_MIN_FILE_AGE_SECONDS:
        result.status = "waiting_stable"
        result.message = (
            f"Plik ma {max(0, int(age_seconds))} s; wymagane {DOCUMASTER_MIN_FILE_AGE_SECONDS} s."
        )
        log(f"Documaster: {source_name}: {result.status}.")
        return result

    payload = smb.read_binary([DOCUMASTER_SMB_DIR, source_name])
    result.file_hash = hashlib.sha256(payload).hexdigest()
    previous = store.get_file(result.file_hash)
    if previous is not None and previous["status"] == "failed_invalid":
        result.status = "failed_unchanged"
        result.message = "Niezmieniony plik byl juz odrzucony jako niepoprawny."
        log(f"Documaster: {source_name}: {result.status}.")
        return result

    try:
        parsed = parse_documaster_file(source_name, payload)
    except DocumasterParseError as exc:
        result.status = "failed_invalid"
        result.message = str(exc)
        log(f"Documaster: {source_name}: {result.status}: {exc}")
        return result

    result.report_customer = parsed.report_customer
    if previous is not None and previous["status"] in {
        "processed",
        "processed_warning",
        "duplicate_file",
    }:
        result.status = "duplicate_file"
        result.message = "Hash pliku byl juz przetworzony."
        if execute:
            result.archive_path = _archive_source(
                smb,
                source_name,
                parsed.report_customer,
                result.file_hash,
                now,
            )
        log(f"Documaster: {source_name}: {result.status}.")
        return result

    try:
        result.rows = firebird.process(
            parsed,
            execute=execute,
            file_hash=result.file_hash,
        )
    except Exception as exc:
        result.status = "failed_retryable"
        result.message = f"{type(exc).__name__}: {exc}"
        log(f"Documaster: {source_name}: {result.status}: {exc}")
        return result

    if not execute:
        result.status = "dry_run_warning" if result.has_warning else "dry_run"
        result.message = "Dry-run zakonczony; plik pozostawiono bez zmian."
        log(f"Documaster: {source_name}: {result.status}, {dict(result.status_counts)}.")
        return result

    result.status = "processed_warning" if result.has_warning else "processed"
    result.archive_path = _archive_source(
        smb,
        source_name,
        parsed.report_customer,
        result.file_hash,
        now,
    )
    result.message = "Import zakonczony i plik zarchiwizowany."
    log(f"Documaster: {source_name}: {result.status}, {dict(result.status_counts)}.")
    return result


def _archive_source(
    smb: SmbClient,
    source_name: str,
    report_customer: str,
    file_hash: str,
    now: dt.datetime,
) -> str:
    entries = smb.list_directory([DOCUMASTER_SMB_DIR])
    folder = _archive_folder(report_customer, entries)
    smb.ensure_directory([DOCUMASTER_SMB_DIR, folder])
    target_parts = [DOCUMASTER_SMB_DIR, folder, source_name]
    if smb.exists(target_parts):
        existing_hash = hashlib.sha256(smb.read_binary(target_parts)).hexdigest()
        if existing_hash == file_hash:
            smb.remove_file([DOCUMASTER_SMB_DIR, source_name])
            return str(PureWindowsPath(DOCUMASTER_SMB_DIR, folder, source_name))
        source_path = Path(source_name)
        archive_name = f"{source_path.stem}_{now:%Y%m%d_%H%M%S}_{file_hash[:8]}{source_path.suffix}"
        target_parts = [DOCUMASTER_SMB_DIR, folder, archive_name]
    return smb.move([DOCUMASTER_SMB_DIR, source_name], target_parts)


def _archive_folder(report_customer: str, entries: list[str]) -> str:
    normalized_customer = report_customer.casefold()
    candidates = [
        name
        for name in entries
        if Path(name).suffix.casefold() not in SUPPORTED_SUFFIXES
        and name.casefold() in normalized_customer
    ]
    if candidates:
        return sorted(candidates, key=lambda item: (len(item), item.casefold()))[0]
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", report_customer).strip(" ._")
    return cleaned[:80] or "nieznany_klient"


def _read_csv_rows(payload: bytes) -> list[list[Any]]:
    text: str
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = payload.decode("cp1250")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    return [list(row) for row in csv.reader(StringIO(text), delimiter=delimiter)]


def _read_xlsx_rows(payload: bytes) -> list[list[Any]]:
    try:
        # Eksport Documaster potrafi deklarowac wymiar A1:A1 mimo danych do AG.
        # Tryb read_only ufa tej deklaracji, dlatego arkusz trzeba zaladowac normalnie.
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style",
                category=UserWarning,
            )
            workbook = load_workbook(BytesIO(payload), read_only=False, data_only=True)
    except Exception as exc:
        raise DocumasterParseError(f"Nie mozna odczytac XLSX: {exc}") from exc
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _parse_report_customer(title: str) -> str:
    match = re.search(r"\bdla:\s*(.+?)\s+od\s+\d{4}[.-]\d{2}[.-]\d{2}", title, re.IGNORECASE)
    if not match:
        raise DocumasterParseError("Nie mozna odczytac klienta z naglowka raportu.")
    return match.group(1).strip()


def _parse_documaster_row(row_number: int, row: dict[str, Any]) -> DocumasterRecord:
    serial = str(row.get("Numer seryjny") or "").strip()
    if not serial:
        raise DocumasterParseError(f"Wiersz {row_number}: brak numeru seryjnego.")
    counter_datetime = _parse_documaster_datetime(
        row.get("Data ostatniego odczytu"),
        row_number,
    )
    total = _required_int(row.get("Licznik główny"), row_number, "Licznik główny")
    mono = _required_int(row.get("Suma Cz/B"), row_number, "Suma Cz/B")
    color = _required_int(row.get("Suma Kolor"), row_number, "Suma Kolor")
    copier_mono = _optional_int(row.get("Kopiarka Cz/B"), row_number, "Kopiarka Cz/B")
    copier_color = _optional_int(row.get("Kopiarka Kolor"), row_number, "Kopiarka Kolor")
    printer_mono = _optional_int(row.get("Drukarka Cz/B"), row_number, "Drukarka Cz/B")
    printer_color = _optional_int(row.get("Drukarka Kolor"), row_number, "Drukarka Kolor")
    scanner_mono = _optional_int(row.get("Skaner Cz/B"), row_number, "Skaner Cz/B")
    scanner_color = _optional_int(row.get("Skaner Kolor"), row_number, "Skaner Kolor")
    return DocumasterRecord(
        row_number=row_number,
        serial=serial,
        counter_datetime=counter_datetime,
        source_model=str(row.get("Model") or "").strip(),
        total=total,
        mono=mono,
        color=color,
        copier_total=_sum_optional(copier_mono, copier_color),
        copier_mono=copier_mono,
        copier_color=copier_color,
        printer_total=_sum_optional(printer_mono, printer_color),
        printer_mono=printer_mono,
        printer_color=printer_color,
        scanner_total=_sum_optional(scanner_mono, scanner_color),
    )


def _parse_documaster_datetime(value: Any, row_number: int) -> dt.datetime:
    if isinstance(value, dt.datetime):
        return value.replace(tzinfo=None)
    text = str(value or "").strip()
    for pattern in ("%Y.%m.%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(text, pattern)
        except ValueError:
            continue
    raise DocumasterParseError(
        f"Wiersz {row_number}: niepoprawna Data ostatniego odczytu: {text!r}."
    )


def _required_int(value: Any, row_number: int, name: str) -> int:
    parsed = _optional_int(value, row_number, name)
    if parsed is None:
        raise DocumasterParseError(f"Wiersz {row_number}: brak wartosci {name}.")
    return parsed


def _optional_int(value: Any, row_number: int, name: str) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        number = float(text)
    except ValueError as exc:
        raise DocumasterParseError(
            f"Wiersz {row_number}: niepoprawna wartosc {name}: {value!r}."
        ) from exc
    if number < 0 or not number.is_integer():
        raise DocumasterParseError(
            f"Wiersz {row_number}: {name} musi byc nieujemna liczba calkowita."
        )
    return int(number)


def _sum_optional(first: int | None, second: int | None) -> int | None:
    if first is None and second is None:
        return None
    return (first or 0) + (second or 0)


def _expected_customer_id(matches: dict[str, list[DeviceMatch]]) -> int | None:
    customer_counts: Counter[int] = Counter()
    for devices in matches.values():
        if len(devices) == 1:
            customer_counts[devices[0].id_klient] += 1
    mapped = sum(customer_counts.values())
    if mapped == 0:
        return None
    most_common = customer_counts.most_common()
    if len(most_common) > 1 and most_common[0][1] == most_common[1][1]:
        return None
    customer_id, count = most_common[0]
    if mapped > 1 and count / mapped < 0.8:
        return None
    return customer_id
